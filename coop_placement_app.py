import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import pulp
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Page config
st.set_page_config(page_title="Co-op Placement Optimizer", layout="wide", page_icon="🎓")

# Initialize session state
if 'data_loaded' not in st.session_state:
    st.session_state.data_loaded = False
if 'students_df' not in st.session_state:
    st.session_state.students_df = None
if 'companies_df' not in st.session_state:
    st.session_state.companies_df = None
if 'rankings_df' not in st.session_state:
    st.session_state.rankings_df = None
if 'wide_format_df' not in st.session_state:
    st.session_state.wide_format_df = None

# Sidebar navigation
st.sidebar.title("🎓 Co-op Placement Optimizer")
page = st.sidebar.radio("Navigation", 
                        ["Data Setup", "Manual Data Editor", "Exploratory Analysis", "Optimization"],
                        index=0)

# ============================================================================
# HELPER FUNCTIONS FOR WIDE FORMAT
# ============================================================================

def create_wide_format_template(n_students=11, n_companies=11):
    """Create Excel template in wide format with students as columns"""
    
    # Default student names
    default_students = ['Aiden', 'Angie', 'George', 'James', 'Josh', 
                       'Kalea', 'Kenzie', 'Prapann', 'Sofia', 'Tony', 'Vihaan']
    
    # Extend or truncate based on n_students
    if n_students <= len(default_students):
        student_names = default_students[:n_students]
    else:
        student_names = default_students + [f'Student_{i+1}' for i in range(len(default_students), n_students)]
    
    # Default companies with groups
    default_companies = [
        ('Group A', 'EY', 1, 1),
        ('', 'Finity', 1, 1),
        ('', 'PwC', 1, 1),
        ('Group B', 'Allianz', 1, 1),
        ('', 'Aware Super', 1, 1),
        ('', 'IAG', 1, 1),
        ('', 'Suncorp', 1, 1),
        ('', 'Toyota Finance', 1, 1),
        ('Group C', 'APRA', 1, 1),
        ('', 'icare', 1, 1),
        ('', 'NDIA', 1, 1),
    ]
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Placement Data"
    
    # Headers
    headers = ['Group', 'Companies', 'IT2 CAP', 'IT3 CAP'] + student_names
    
    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Fill in company data
    for row_idx, (group, company, it2_cap, it3_cap) in enumerate(default_companies[:n_companies], 2):
        ws.cell(row_idx, 1, group)
        ws.cell(row_idx, 2, company)
        ws.cell(row_idx, 3, it2_cap)
        ws.cell(row_idx, 4, it3_cap)
        
        # Initialize ranking cells with empty values
        for col_idx in range(5, 5 + len(student_names)):
            cell = ws.cell(row_idx, col_idx)
            cell.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    for col_idx in range(5, 5 + len(student_names)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10
    
    # Add instructions in a separate area
    instructions_row = n_companies + 3
    ws.cell(instructions_row, 1, "INSTRUCTIONS:").font = Font(bold=True, size=12)
    ws.cell(instructions_row + 1, 1, "1. Fill in Group names (Group A, Group B, etc.) for industry categories")
    ws.cell(instructions_row + 2, 1, "2. Enter company names in the 'Companies' column")
    ws.cell(instructions_row + 3, 1, "3. Set IT2 CAP and IT3 CAP (0 if not offering)")
    ws.cell(instructions_row + 4, 1, "4. Add/modify student names in column headers")
    ws.cell(instructions_row + 5, 1, "5. Enter rankings (1-11, where 1 = most preferred, 11 = least preferred)")
    ws.cell(instructions_row + 6, 1, "6. Save and upload to the app")
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def parse_wide_format_excel(file):
    """Parse uploaded Excel file in wide format"""
    try:
        # Read the Excel file
        df = pd.read_excel(file, sheet_name=0)
        
        # Validate required columns
        required_cols = ['Group', 'Companies', 'IT2 CAP', 'IT3 CAP']
        if not all(col in df.columns for col in required_cols):
            return None, None, None, None, "Missing required columns: Group, Companies, IT2 CAP, IT3 CAP"
        
        # Forward fill the Group column
        df['Group'] = df['Group'].fillna(method='ffill')
        
        # Clean up: remove rows where Companies is NaN
        df = df.dropna(subset=['Companies'])
        
        # Extract student names (all columns after IT3 CAP)
        student_cols = [col for col in df.columns if col not in required_cols]
        
        if len(student_cols) == 0:
            return None, None, None, None, "No student columns found"
        
        # Create students dataframe
        students_df = pd.DataFrame({
            'student_id': range(1, len(student_cols) + 1),
            'student_name': student_cols
        })
        
        # Create companies dataframe
        companies_df = pd.DataFrame({
            'company_id': range(1, len(df) + 1),
            'company_name': df['Companies'].values,
            'industry': df['Group'].values,
            'it2_capacity': df['IT2 CAP'].fillna(0).astype(int).values,
            'it3_capacity': df['IT3 CAP'].fillna(0).astype(int).values
        })
        
        # Create rankings dataframe (long format)
        rankings_data = []
        for company_idx, company_id in enumerate(companies_df['company_id']):
            for student_idx, student_col in enumerate(student_cols):
                student_id = student_idx + 1
                ranking = df[student_col].iloc[company_idx]
                
                # Handle NaN rankings
                if pd.isna(ranking):
                    ranking = 0
                else:
                    ranking = int(ranking)
                
                rankings_data.append({
                    'student_id': student_id,
                    'company_id': company_id,
                    'ranking': ranking
                })
        
        rankings_df = pd.DataFrame(rankings_data)
        
        # Store the wide format for editing
        wide_format_df = df.copy()
        
        return students_df, companies_df, rankings_df, wide_format_df, None
        
    except Exception as e:
        return None, None, None, None, f"Error parsing file: {str(e)}"

def convert_to_wide_format(students_df, companies_df, rankings_df):
    """Convert internal format back to wide format for editing"""
    
    # Create base dataframe with company info
    wide_df = companies_df[['company_name', 'industry', 'it2_capacity', 'it3_capacity']].copy()
    wide_df = wide_df.rename(columns={
        'company_name': 'Companies',
        'industry': 'Group',
        'it2_capacity': 'IT2 CAP',
        'it3_capacity': 'IT3 CAP'
    })
    
    # Add student columns
    for _, student in students_df.iterrows():
        student_id = student['student_id']
        student_name = student['student_name']
        
        # Get rankings for this student
        student_rankings = rankings_df[rankings_df['student_id'] == student_id].set_index('company_id')['ranking']
        
        # Add column with rankings (in company order)
        wide_df[student_name] = companies_df['company_id'].map(student_rankings).fillna(0).astype(int)
    
    # Optimize Group column display (show group name only on first occurrence)
    current_group = None
    for idx in range(len(wide_df)):
        if wide_df.iloc[idx]['Group'] == current_group:
            wide_df.iloc[idx, wide_df.columns.get_loc('Group')] = ''
        else:
            current_group = wide_df.iloc[idx]['Group']
    
    return wide_df

def convert_from_wide_format(wide_df):
    """Convert edited wide format back to internal format"""
    
    # Forward fill Group column
    df = wide_df.copy()
    df['Group'] = df['Group'].replace('', np.nan).fillna(method='ffill')
    
    # Extract student columns
    student_cols = [col for col in df.columns if col not in ['Group', 'Companies', 'IT2 CAP', 'IT3 CAP']]
    
    # Create students dataframe
    students_df = pd.DataFrame({
        'student_id': range(1, len(student_cols) + 1),
        'student_name': student_cols
    })
    
    # Create companies dataframe
    companies_df = pd.DataFrame({
        'company_id': range(1, len(df) + 1),
        'company_name': df['Companies'].values,
        'industry': df['Group'].values,
        'it2_capacity': df['IT2 CAP'].fillna(0).astype(int).values,
        'it3_capacity': df['IT3 CAP'].fillna(0).astype(int).values
    })
    
    # Create rankings dataframe
    rankings_data = []
    for company_idx, company_id in enumerate(companies_df['company_id']):
        for student_idx, student_col in enumerate(student_cols):
            student_id = student_idx + 1
            ranking = df[student_col].iloc[company_idx]
            
            if pd.isna(ranking):
                ranking = 0
            else:
                ranking = int(ranking)
            
            rankings_data.append({
                'student_id': student_id,
                'company_id': company_id,
                'ranking': ranking
            })
    
    rankings_df = pd.DataFrame(rankings_data)
    
    return students_df, companies_df, rankings_df

# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def check_optimization_feasibility(students_df, companies_df):
    """Check if the optimization problem is likely feasible"""
    errors = []
    warnings = []
    
    n_students = len(students_df)
    
    # Check total capacities (HARD CONSTRAINT)
    total_it2_cap = companies_df['it2_capacity'].sum()
    total_it3_cap = companies_df['it3_capacity'].sum()
    
    if total_it2_cap < n_students:
        errors.append(f"❌ Total IT2 capacity ({total_it2_cap}) < number of students ({n_students})")
        errors.append(f"   → Related to Constraint 1: Σⱼ xᵢⱼ = 1 (every student needs one IT2)")
    
    if total_it3_cap < n_students:
        errors.append(f"❌ Total IT3 capacity ({total_it3_cap}) < number of students ({n_students})")
        errors.append(f"   → Related to Constraint 2: Σⱼ yᵢⱼ = 1 (every student needs one IT3)")
    
    # Check industry capacities (WORST-CASE WARNING, not hard error)
    industry_caps = companies_df.groupby('industry')[['it2_capacity', 'it3_capacity']].sum()
    
    for industry, row in industry_caps.iterrows():
        total_industry_cap = row['it2_capacity'] + row['it3_capacity']
        if total_industry_cap < n_students:
            warnings.append(
                f"⚠️ Industry '{industry}' total capacity ({int(total_industry_cap)}) < students ({n_students})\n"
                f"   → Related to Constraint 6: Σⱼ∈Industry (xᵢⱼ + yᵢⱼ) ≤ 1 (at most one placement per industry)\n"
                f"   → This is a WORST-CASE check: if all students prefer this industry, optimization will fail\n"
                f"   → However, if preferences are distributed, optimization may still succeed\n"
                f"   → You can proceed, but be aware the problem might be infeasible"
            )
    
    # Warnings for tight capacities
    if total_it2_cap < n_students * 1.5:
        warnings.append(f"⚠️ IT2 capacity is tight ({total_it2_cap} vs {n_students} students) - solution quality may be limited")
    
    if total_it3_cap < n_students * 1.5:
        warnings.append(f"⚠️ IT3 capacity is tight ({total_it3_cap} vs {n_students} students) - solution quality may be limited")
    
    is_feasible = len(errors) == 0
    return is_feasible, errors, warnings

# ============================================================================
# PAGE 1: DATA SETUP
# ============================================================================

if page == "Data Setup":
    st.title("📊 Data Setup")
    
    st.write("""
    Upload your placement data in the **wide format** (students as columns) or download a template to get started.
    """)
    
    # Template download section
    st.header("📥 Download Template")
    
    col1, col2 = st.columns(2)
    with col1:
        template_students = st.number_input("Number of Students", min_value=1, max_value=50, value=11)
    with col2:
        template_companies = st.number_input("Number of Companies", min_value=1, max_value=50, value=11)
    
    if st.button("Generate Template"):
        template_buffer = create_wide_format_template(template_students, template_companies)
        st.download_button(
            label="📥 Download Excel Template",
            data=template_buffer,
            file_name="placement_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    st.markdown("---")
    
    # File upload section
    st.header("📤 Upload Data")
    
    uploaded_file = st.file_uploader("Upload Excel file (wide format)", type=['xlsx'])
    
    if uploaded_file is not None:
        with st.spinner("Processing uploaded file..."):
            students_df, companies_df, rankings_df, wide_format_df, error = parse_wide_format_excel(uploaded_file)
            
            if error:
                st.error(f"Error: {error}")
            else:
                st.success("✅ Data loaded successfully!")
                
                # Store in session state
                st.session_state.students_df = students_df
                st.session_state.companies_df = companies_df
                st.session_state.rankings_df = rankings_df
                st.session_state.wide_format_df = wide_format_df
                st.session_state.data_loaded = True
                
                # Show summary
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Students", len(students_df))
                with col2:
                    st.metric("Companies", len(companies_df))
                with col3:
                    st.metric("Industries", companies_df['industry'].nunique())
                
                # Preview data
                with st.expander("📋 Preview Loaded Data"):
                    st.subheader("Wide Format View")
                    st.dataframe(wide_format_df, use_container_width=True, height=400)
    
    # Show current data status
    if st.session_state.data_loaded:
        st.markdown("---")
        st.success("✅ Data is loaded and ready for analysis!")
        
        if st.button("🗑️ Clear Data"):
            st.session_state.data_loaded = False
            st.session_state.students_df = None
            st.session_state.companies_df = None
            st.session_state.rankings_df = None
            st.session_state.wide_format_df = None
            st.experimental_rerun()

# ============================================================================
# PAGE 2: MANUAL DATA EDITOR
# ============================================================================

elif page == "Manual Data Editor":
    st.title("✏️ Manual Data Editor")
    
    if not st.session_state.data_loaded:
        st.warning("⚠️ Please load data first in the Data Setup page.")
    else:
        st.write("""
        Edit your placement data directly in the wide format. Changes will be automatically converted back to the internal format.
        """)
        
        # Convert current data to wide format if not already available
        if st.session_state.wide_format_df is None:
            st.session_state.wide_format_df = convert_to_wide_format(
                st.session_state.students_df,
                st.session_state.companies_df,
                st.session_state.rankings_df
            )
        
        # Edit the wide format dataframe
        st.subheader("Edit Data (Wide Format)")
        st.info("💡 You can add/remove rows, change capacities, modify rankings, etc.")
        
        edited_df = st.data_editor(
            st.session_state.wide_format_df,
            use_container_width=True,
            num_rows="dynamic",
            height=600,
            column_config={
                "Group": st.column_config.TextColumn("Group", width="medium"),
                "Companies": st.column_config.TextColumn("Companies", width="large"),
                "IT2 CAP": st.column_config.NumberColumn("IT2 CAP", min_value=0, max_value=20, width="small"),
                "IT3 CAP": st.column_config.NumberColumn("IT3 CAP", min_value=0, max_value=20, width="small"),
            }
        )
        
        if st.button("💾 Save Changes", type="primary"):
            with st.spinner("Converting and saving changes..."):
                try:
                    # Convert back to internal format
                    students_df, companies_df, rankings_df = convert_from_wide_format(edited_df)
                    
                    # Update session state
                    st.session_state.students_df = students_df
                    st.session_state.companies_df = companies_df
                    st.session_state.rankings_df = rankings_df
                    st.session_state.wide_format_df = edited_df
                    
                    st.success("✅ Changes saved successfully!")
                    
                    # Show summary
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Students", len(students_df))
                    with col2:
                        st.metric("Companies", len(companies_df))
                    with col3:
                        st.metric("Industries", companies_df['industry'].nunique())
                        
                except Exception as e:
                    st.error(f"Error saving changes: {str(e)}")

# ============================================================================
# PAGE 3: EXPLORATORY ANALYSIS
# ============================================================================

elif page == "Exploratory Analysis":
    st.title("📊 Exploratory Analysis")
    
    if not st.session_state.data_loaded:
        st.warning("⚠️ Please load data first in the Data Setup page.")
    else:
        students_df = st.session_state.students_df
        companies_df = st.session_state.companies_df
        rankings_df = st.session_state.rankings_df
        
        # Summary statistics
        st.header("Summary Statistics")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Students", len(students_df))
        with col2:
            st.metric("Total Companies", len(companies_df))
        with col3:
            st.metric("Total Rankings", len(rankings_df[rankings_df['ranking'] > 0]))
        with col4:
            st.metric("Avg Ranking", f"{rankings_df[rankings_df['ranking'] > 0]['ranking'].mean():.2f}")
        
        # Company analysis
        st.header("Company Analysis")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Capacity by Industry")
            industry_capacity = companies_df.groupby('industry')[['it2_capacity', 'it3_capacity']].sum()
            fig = px.bar(industry_capacity, barmode='group', title="IT2 and IT3 Capacity by Industry")
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            st.subheader("Companies per Industry")
            industry_counts = companies_df['industry'].value_counts()
            fig = px.pie(values=industry_counts.values, names=industry_counts.index, 
                        title="Distribution of Companies by Industry")
            st.plotly_chart(fig, use_container_width=True)
        
        # Student preferences heatmap
        st.header("Student-Company Preference Heatmap")
        
        # Create pivot table for heatmap
        heatmap_data = rankings_df.pivot(index='student_id', columns='company_id', values='ranking')
        
        # Map student and company names
        student_names = students_df.set_index('student_id')['student_name']
        company_names = companies_df.set_index('company_id')['company_name']
        
        heatmap_data.index = heatmap_data.index.map(student_names)
        heatmap_data.columns = heatmap_data.columns.map(company_names)
        
        fig = px.imshow(heatmap_data, 
                       labels=dict(x="Company", y="Student", color="Ranking"),
                       title="Student Rankings Heatmap (1=Best, 11=Worst)",
                       aspect="auto",
                       color_continuous_scale='RdYlGn_r')  # Reversed: green for 1 (best), red for 11 (worst)
        fig.update_xaxes(side="top", tickangle=-45)
        st.plotly_chart(fig, use_container_width=True)

# ============================================================================
# PAGE 4: OPTIMIZATION
# ============================================================================

elif page == "Optimization":
    st.title("⚡ Placement Optimization")
    
    if not st.session_state.data_loaded:
        st.warning("⚠️ Please load data first in the Data Setup page.")
    else:
        st.write("""
        This page solves the integer linear programming problem to optimally assign students 
        to companies for IT2 and IT3 placements while respecting all constraints.
        """)
        
        students_df = st.session_state.students_df
        companies_df = st.session_state.companies_df
        rankings_df = st.session_state.rankings_df
        
        n_students = len(students_df)
        n_companies = len(companies_df)
        n_variables = n_students * n_companies * 2
        
        st.info(f"**Problem Size:** {n_students} students × {n_companies} companies = {n_variables} decision variables")
        
        if n_variables > 10000:
            st.warning("⚠️ Large problem size. Optimization may take several minutes.")
        
        # PRE-OPTIMIZATION VALIDATION
        st.header("Pre-Optimization Checks")
        
        # Run feasibility checks
        is_feasible, feasibility_errors, feasibility_warnings = check_optimization_feasibility(students_df, companies_df)
        
        # Display validation results
        if feasibility_errors:
            st.error("🚫 **Critical Issues Detected:**")
            for err in feasibility_errors:
                st.write(err)
            st.error("**Cannot proceed with optimization. Please fix the issues above.**")
            st.stop()
        else:
            st.success("✅ Basic feasibility checks passed!")
        
        if feasibility_warnings:
            with st.expander("⚠️ View Warnings (Click to expand)", expanded=False):
                for warn in feasibility_warnings:
                    st.write(warn)
                st.info("💡 These warnings won't prevent optimization but may affect solution quality.")
        
        # Display industry capacity analysis
        st.subheader("Industry Capacity Analysis")
        industry_summary = companies_df.groupby('industry').agg({
            'it2_capacity': 'sum',
            'it3_capacity': 'sum',
            'company_id': 'count'
        }).rename(columns={'company_id': 'num_companies'})
        industry_summary['total_capacity'] = industry_summary['it2_capacity'] + industry_summary['it3_capacity']
        industry_summary = industry_summary.sort_values('total_capacity', ascending=False)
        
        st.dataframe(industry_summary, use_container_width=True)
        
        if st.button("🚀 Solve Optimization Problem", type="primary"):
            with st.spinner("Solving optimization problem..."):
                # Create the optimization problem
                prob = pulp.LpProblem("CoopPlacement", pulp.LpMaximize)
                
                students = students_df['student_id'].tolist()
                companies = companies_df['company_id'].tolist()
                
                # Decision variables
                x = {}  # IT2 assignments
                y = {}  # IT3 assignments
                
                for i in students:
                    for j in companies:
                        x[i, j] = pulp.LpVariable(f"x_{i}_{j}", cat='Binary')
                        y[i, j] = pulp.LpVariable(f"y_{i}_{j}", cat='Binary')
                
                # Handle missing rankings by setting them to 0
                rankings_dict = rankings_df.set_index(['student_id', 'company_id'])['ranking'].to_dict()
                
                # Ensure all student-company pairs have a ranking (default to 0 if missing)
                complete_rankings_dict = {}
                for i in students:
                    for j in companies:
                        complete_rankings_dict[i, j] = rankings_dict.get((i, j), 0)
                
                # Objective function: Minimize total ranking (1 is best, 11 is worst)
                # Equivalent to maximizing (12 - ranking)
                prob += pulp.lpSum([(12 - complete_rankings_dict[i, j]) * (x[i, j] + y[i, j]) 
                                   for i in students for j in companies])
                
                # Constraint 1: Every student does exactly one IT2 placement
                for i in students:
                    prob += pulp.lpSum([x[i, j] for j in companies]) == 1, f"IT2_assignment_student_{i}"
                
                # Constraint 2: Every student does exactly one IT3 placement
                for i in students:
                    prob += pulp.lpSum([y[i, j] for j in companies]) == 1, f"IT3_assignment_student_{i}"
                
                # Constraint 3: IT2 capacity constraints
                for j in companies:
                    capacity = companies_df[companies_df['company_id'] == j]['it2_capacity'].values[0]
                    prob += pulp.lpSum([x[i, j] for i in students]) <= capacity, f"IT2_capacity_company_{j}"
                
                # Constraint 4: IT3 capacity constraints
                for j in companies:
                    capacity = companies_df[companies_df['company_id'] == j]['it3_capacity'].values[0]
                    prob += pulp.lpSum([y[i, j] for i in students]) <= capacity, f"IT3_capacity_company_{j}"
                
                # Constraint 5: Students get different companies for IT2 and IT3
                for i in students:
                    for j in companies:
                        prob += x[i, j] + y[i, j] <= 1, f"Different_companies_student_{i}_company_{j}"
                
                # Constraint 6: Industry diversity - at most one placement per industry per student
                industry_companies = companies_df.groupby('industry')['company_id'].apply(list).to_dict()
                
                for i in students:
                    for industry, company_list in industry_companies.items():
                        prob += pulp.lpSum([x[i, j] + y[i, j] for j in company_list]) <= 1, \
                                f"Industry_diversity_student_{i}_industry_{industry}"
                
                # Solve the problem
                solver = pulp.PULP_CBC_CMD(msg=0)
                prob.solve(solver)
                
                # Display results
                status_text = pulp.LpStatus[prob.status]
                
                if prob.status == 1:  # Optimal solution found
                    st.success(f"✅ Optimization Status: {status_text}")
                else:
                    st.error(f"❌ Optimization Status: {status_text}")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Objective Value", f"{pulp.value(prob.objective):.2f}")
                with col2:
                    max_possible = rankings_df['ranking'].max() * 2 * len(students)
                    st.metric("Max Possible", f"{max_possible}")
                with col3:
                    if pulp.value(prob.objective) and max_possible > 0:
                        efficiency = (pulp.value(prob.objective) / max_possible) * 100
                        st.metric("Efficiency", f"{efficiency:.1f}%")
                
                if prob.status == 1:
                    # Extract assignments
                    it2_assignments = []
                    it3_assignments = []
                    
                    for i in students:
                        for j in companies:
                            if pulp.value(x[i, j]) == 1:
                                student_name = students_df[students_df['student_id'] == i]['student_name'].values[0]
                                company_name = companies_df[companies_df['company_id'] == j]['company_name'].values[0]
                                industry = companies_df[companies_df['company_id'] == j]['industry'].values[0]
                                ranking = complete_rankings_dict[i, j]
                                it2_assignments.append({
                                    'Student': student_name,
                                    'Company': company_name,
                                    'Industry': industry,
                                    'Ranking': ranking
                                })
                            
                            if pulp.value(y[i, j]) == 1:
                                student_name = students_df[students_df['student_id'] == i]['student_name'].values[0]
                                company_name = companies_df[companies_df['company_id'] == j]['company_name'].values[0]
                                industry = companies_df[companies_df['company_id'] == j]['industry'].values[0]
                                ranking = complete_rankings_dict[i, j]
                                it3_assignments.append({
                                    'Student': student_name,
                                    'Company': company_name,
                                    'Industry': industry,
                                    'Ranking': ranking
                                })
                    
                    it2_df = pd.DataFrame(it2_assignments)
                    it3_df = pd.DataFrame(it3_assignments)
                    
                    st.header("Placement Assignments")
                    
                    tab1, tab2, tab3 = st.tabs(["IT2 Placements", "IT3 Placements", "Combined View"])
                    
                    with tab1:
                        st.subheader("IT2 Placements")
                        st.dataframe(it2_df.sort_values('Student'), hide_index=True, height=500)
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Average IT2 Ranking", f"{it2_df['Ranking'].mean():.2f}")
                        with col2:
                            st.metric("Min IT2 Ranking", f"{it2_df['Ranking'].min():.0f}")
                        with col3:
                            st.metric("Max IT2 Ranking", f"{it2_df['Ranking'].max():.0f}")
                    
                    with tab2:
                        st.subheader("IT3 Placements")
                        st.dataframe(it3_df.sort_values('Student'), hide_index=True, height=500)
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Average IT3 Ranking", f"{it3_df['Ranking'].mean():.2f}")
                        with col2:
                            st.metric("Min IT3 Ranking", f"{it3_df['Ranking'].min():.0f}")
                        with col3:
                            st.metric("Max IT3 Ranking", f"{it3_df['Ranking'].max():.0f}")
                    
                    with tab3:
                        combined = it2_df.merge(it3_df, on='Student', suffixes=('_IT2', '_IT3'))
                        st.dataframe(combined, hide_index=True, height=500)
                    
                    st.header("Export Results")
                    
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        it2_df.to_excel(writer, sheet_name='IT2_Placements', index=False)
                        it3_df.to_excel(writer, sheet_name='IT3_Placements', index=False)
                        combined.to_excel(writer, sheet_name='Combined', index=False)
                        
                        # Add summary sheet
                        summary_data = {
                            'Metric': [
                                'Total Students',
                                'Total Companies',
                                'Objective Value',
                                'Max Possible',
                                'Efficiency (%)',
                                'Avg IT2 Ranking',
                                'Avg IT3 Ranking',
                                'Overall Avg Ranking'
                            ],
                            'Value': [
                                len(students),
                                len(companies),
                                f"{pulp.value(prob.objective):.2f}",
                                max_possible,
                                f"{efficiency:.1f}" if 'efficiency' in locals() else 'N/A',
                                f"{it2_df['Ranking'].mean():.2f}",
                                f"{it3_df['Ranking'].mean():.2f}",
                                f"{pd.concat([it2_df['Ranking'], it3_df['Ranking']]).mean():.2f}"
                            ]
                        }
                        summary_df = pd.DataFrame(summary_data)
                        summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    st.download_button(
                        label="📥 Download Results (Excel)",
                        data=output.getvalue(),
                        file_name="placement_results.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error("❌ Optimization failed to find a solution.")
                    
                    st.subheader("Possible Reasons (with constraint references):")
                    reasons = []
                    
                    # Check capacity issues
                    if companies_df['it2_capacity'].sum() < n_students:
                        reasons.append(
                            "• Total IT2 capacity is less than number of students\n"
                            "  → Violates Constraint 1: Σⱼ xᵢⱼ = 1 (every student needs one IT2)"
                        )
                    if companies_df['it3_capacity'].sum() < n_students:
                        reasons.append(
                            "• Total IT3 capacity is less than number of students\n"
                            "  → Violates Constraint 2: Σⱼ yᵢⱼ = 1 (every student needs one IT3)"
                        )
                    
                    # Check industry diversity feasibility
                    industry_capacities = companies_df.groupby('industry')[['it2_capacity', 'it3_capacity']].sum()
                    for industry, row in industry_capacities.iterrows():
                        total_cap = row['it2_capacity'] + row['it3_capacity']
                        if total_cap < n_students:
                            reasons.append(
                                f"• Industry '{industry}' has insufficient total capacity ({int(total_cap)} < {n_students})\n"
                                f"  → Related to Constraint 6: Σⱼ∈Industry (xᵢⱼ + yᵢⱼ) ≤ 1\n"
                                f"  → Student preferences concentrated in this industry beyond its capacity"
                            )
                    
                    if not reasons:
                        reasons.append(
                            "• Unknown infeasibility - the combination of constraints cannot be satisfied\n"
                            "  → Check: Constraint 5 (different companies) + Constraint 6 (industry diversity)\n"
                            "  → May need to adjust capacities, add companies, or review student preferences"
                        )
                    
                    for reason in reasons:
                        st.write(reason)
                    
                    st.info(
                        "💡 **How to fix:**\n\n"
                        "1. Increase company capacities (especially in popular industries)\n"
                        "2. Add more companies to industries with low capacity\n"
                        "3. Check that each industry total capacity is close to the number of students\n"
                        "4. Review student rankings - are they too concentrated in few companies?\n\n"
                        "**Mathematical requirement:** For N students across K industries, optimal distribution "
                        "is when each industry has total capacity ≈ N to avoid infeasibility."
                    )

# ============================================================================
# FOOTER
# ============================================================================

st.sidebar.markdown("---")
st.sidebar.info(f"""
**Co-op Placement Optimizer**  
Version 5.1 - Streamlined Edition
Built with Streamlit & PuLP

Current Data:
- Students: {len(st.session_state.students_df) if st.session_state.data_loaded else 0}
- Companies: {len(st.session_state.companies_df) if st.session_state.data_loaded else 0}
- Industries: {st.session_state.companies_df['industry'].nunique() if st.session_state.data_loaded and len(st.session_state.companies_df) > 0 else 0}

Features:
✅ Wide format Excel (students as columns)
✅ Manual data editing
✅ Smart validation (errors vs warnings)
✅ Constraint-referenced diagnostics
✅ Streamlined optimization results
""")
